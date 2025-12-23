import logging
import os
import tempfile
import urllib.parse
import json
from typing import Any, List, Dict, Optional

from mcp.server.fastmcp import FastMCP

try:
    from dotenv import load_dotenv
except Exception:
    load_dotenv = None

import oss2

# Import exceptions
from excel_mcp.exceptions import (
    ValidationError,
    WorkbookError,
    SheetError,
    DataError,
    FormattingError,
    CalculationError,
    PivotError,
    ChartError
)

# Import from excel_mcp package with consistent _impl suffixes
from excel_mcp.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
    validate_range_in_sheet_operation as validate_range_impl
)
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.workbook import get_workbook_info
from excel_mcp.data import write_data
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_sheet,
    delete_sheet,
    rename_sheet,
    merge_range,
    unmerge_range,
    get_merged_ranges,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
)

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "excel-mcp.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")

# Initialize FastMCP server
mcp = FastMCP(
    "excel-mcp",
    host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
    port=int(os.environ.get("FASTMCP_PORT", "8017")),
    instructions="Excel MCP Server for manipulating Excel files"
)

if load_dotenv is not None:
    # Load env vars from .env if present (useful when running as a service)
    load_dotenv(override=False)

OSS_ACCESS_KEY_ID = os.environ.get("OSS_ACCESS_KEY_ID") or os.environ.get("ACCESS_KEY_ID")
OSS_ACCESS_KEY_SECRET = os.environ.get("OSS_ACCESS_KEY_SECRET") or os.environ.get("ACCESS_KEY_SECRET")
OSS_BUCKET_NAME = os.environ.get("OSS_BUCKET_NAME") or os.environ.get("BUCKET_NAME")
OSS_ENDPOINT = os.environ.get("OSS_ENDPOINT") or os.environ.get("ENDPOINT")

# Optional prefix to prepend to any derived oss_key.
# Example: "tenant/user/thread" so that "UPLOADED_DOCUMENTS/a.xlsx" maps to
# "tenant/user/thread/UPLOADED_DOCUMENTS/a.xlsx".
OSS_KEY_PREFIX = os.environ.get("OSS_KEY_PREFIX", "").strip()

# Virtual FS selector. If enabled, paths like ../UPLOADED_DOCUMENTS/<file> are treated as OSS keys.
OSS_VIRTUAL_UPLOADED_DOCS = os.environ.get("OSS_VIRTUAL_UPLOADED_DOCS", "true").lower() in ("1", "true", "yes", "y")


def _get_oss_bucket() -> oss2.Bucket:
    """Create OSS bucket client lazily so server can start even if OSS isn't configured."""
    if not (OSS_ENDPOINT and OSS_BUCKET_NAME and OSS_ACCESS_KEY_ID and OSS_ACCESS_KEY_SECRET):
        raise ValueError(
            "OSS is not configured. Set ENDPOINT, BUCKET_NAME, ACCESS_KEY_ID, ACCESS_KEY_SECRET"
        )
    auth = oss2.Auth(OSS_ACCESS_KEY_ID, OSS_ACCESS_KEY_SECRET)
    return oss2.Bucket(auth, OSS_ENDPOINT, OSS_BUCKET_NAME)


def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.

    Args:
        filename: Name of Excel file

    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(f"Invalid filename: {filename}, must be an absolute path when not in SSE mode")

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(EXCEL_FILES_PATH, filename)


def _safe_local_filename_from_oss_key(oss_key: str) -> str:
    name = os.path.basename(oss_key)
    return name or "workbook.xlsx"


def _download_oss_object_to_temp(oss_key: str) -> str:
    """Download an OSS object into a unique temp directory and return the local path."""
    bucket = _get_oss_bucket()
    tmp_dir = tempfile.mkdtemp(prefix="excel_mcp_oss_")
    local_path = os.path.join(tmp_dir, _safe_local_filename_from_oss_key(oss_key))
    bucket.get_object_to_file(oss_key, local_path)
    return local_path


def _normalize_relpath_for_oss(path: str) -> str:
    p = (path or "").replace("\\", "/")
    while p.startswith("./"):
        p = p[2:]
    while p.startswith("../"):
        p = p[3:]
    p = p.lstrip("/")
    return p


def _is_virtual_oss_path(filepath: str) -> bool:
    if not OSS_VIRTUAL_UPLOADED_DOCS:
        return False
    if not filepath or os.path.isabs(filepath):
        return False
    p = _normalize_relpath_for_oss(filepath)
    return p.startswith("UPLOADED_DOCUMENTS/")


def _virtual_path_to_oss_key(filepath: str) -> str:
    p = _normalize_relpath_for_oss(filepath)
    if OSS_KEY_PREFIX:
        return f"{OSS_KEY_PREFIX.rstrip('/')}/{p}"
    return p


def _resolve_file_for_operation(filepath: str, allow_create: bool = False) -> tuple[str, Optional[str]]:
    """Resolve filepath into a local path, optionally backed by OSS.

    Returns:
        (local_path, oss_key_or_none)
    """
    if _is_virtual_oss_path(filepath):
        oss_key = _virtual_path_to_oss_key(filepath)
        try:
            local_path = _download_oss_object_to_temp(oss_key)
            return local_path, oss_key
        except Exception:
            if not allow_create:
                raise
            tmp_dir = tempfile.mkdtemp(prefix="excel_mcp_oss_")
            local_path = os.path.join(tmp_dir, _safe_local_filename_from_oss_key(oss_key))
            from excel_mcp.workbook import create_workbook as create_workbook_impl
            create_workbook_impl(local_path)
            return local_path, oss_key

    return get_excel_path(filepath), None


def _maybe_upload_after_mutation(local_path: str, oss_key: Optional[str]) -> None:
    if oss_key:
        _upload_temp_file_to_oss(local_path, oss_key)


def _upload_temp_file_to_oss(local_path: str, oss_key: str) -> Dict[str, Any]:
    """Upload local file back to OSS and return useful metadata."""
    bucket = _get_oss_bucket()
    with open(local_path, "rb") as f:
        bucket.put_object(oss_key, f)

    filename = os.path.basename(local_path)
    params = {"response-content-disposition": f'attachment; filename="{filename}"'}
    signed_url = bucket.sign_url("GET", oss_key, 86400, params=params)

    try:
        object_url = bucket.object_url(oss_key)
    except Exception:
        endpoint = str(OSS_ENDPOINT or "").replace("https://", "").replace("http://", "").rstrip("/")
        object_url = f"https://{OSS_BUCKET_NAME}.{endpoint}/{urllib.parse.quote(oss_key)}" if endpoint and OSS_BUCKET_NAME else None

    try:
        head = bucket.head_object(oss_key)
        etag = getattr(head, "etag", None)
        last_modified = getattr(head, "last_modified", None)
        content_length = getattr(head, "content_length", None)
    except Exception:
        etag = None
        last_modified = None
        content_length = None

    return {
        "oss_key": oss_key,
        "oss_url": signed_url,
        "oss_object_url": object_url,
        "oss_etag": etag,
        "oss_last_modified": last_modified,
        "oss_content_length": content_length,
        "file_name": filename,
    }


@mcp.tool()
def oss_get_workbook_metadata(oss_key: str, include_ranges: bool = False) -> str:
    """Get workbook metadata for an OSS-backed Excel file (download -> inspect)."""
    try:
        local_path = _download_oss_object_to_temp(oss_key)
        info = get_workbook_info(local_path, include_ranges=include_ranges)
        info["oss_key"] = oss_key
        return json.dumps(info, default=str)
    except Exception as e:
        logger.error(f"Error in oss_get_workbook_metadata: {e}")
        raise


@mcp.tool()
def oss_read_data_from_excel(
    oss_key: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False,
) -> str:
    """Read range data from an OSS-backed Excel file (download -> read)."""
    try:
        local_path = _download_oss_object_to_temp(oss_key)
        return read_data_from_excel(
            filepath=local_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,
            preview_only=preview_only,
        )
    except Exception as e:
        logger.error(f"Error in oss_read_data_from_excel: {e}")
        raise


@mcp.tool()
def oss_write_data_to_excel(
    oss_key: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """Write data to an OSS-backed Excel file (download -> write -> upload)."""
    try:
        try:
            local_path = _download_oss_object_to_temp(oss_key)
        except Exception:
            # If object doesn't exist (or download fails), create a new workbook locally.
            tmp_dir = tempfile.mkdtemp(prefix="excel_mcp_oss_")
            local_path = os.path.join(tmp_dir, _safe_local_filename_from_oss_key(oss_key))
            from excel_mcp.workbook import create_workbook as create_workbook_impl
            create_workbook_impl(local_path)

        msg = write_data_to_excel(
            filepath=local_path,
            sheet_name=sheet_name,
            data=data,
            start_cell=start_cell,
        )

        upload_meta = _upload_temp_file_to_oss(local_path, oss_key)
        return json.dumps({"message": msg, **upload_meta}, default=str)
    except Exception as e:
        logger.error(f"Error in oss_write_data_to_excel: {e}")
        raise


@mcp.tool()
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """
    Apply Excel formula to cell.
    Excel formula will write to cell with verification.
    """
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)

        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"

        # If valid, apply the formula
        from excel_mcp.calculations import apply_formula as apply_formula_impl
        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise


@mcp.tool()
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it."""
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise


@mcp.tool()
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None
) -> str:
    """Apply formatting to a range of cells."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        from excel_mcp.formatting import format_range as format_range_func

        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format  # This can be None
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise


@mcp.tool()
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False
) -> str:
    """
    Read data from Excel worksheet with cell metadata including validation rules.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only

    Returns:
        JSON string containing structured cell data with validation metadata.
        Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        from excel_mcp.data import read_excel_range_with_metadata
        result = read_excel_range_with_metadata(
            full_path,
            sheet_name,
            start_cell,
            end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"

        # Return as formatted JSON string
        import json
        return json.dumps(result, indent=2, default=str)

    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise


@mcp.tool()
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """
    Write data to Excel worksheet.
    Excel formula will write to cell without any verification.

    PARAMETERS:
        filepath: Path to Excel file
        sheet_name: Name of worksheet to write to
        data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
        start_cell: Cell to start writing to, default is "A1"
    """
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath, allow_create=True)
        result = write_data(full_path, sheet_name, data, start_cell)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise


@mcp.tool()
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath, allow_create=True)
        from excel_mcp.workbook import create_workbook as create_workbook_impl
        create_workbook_impl(full_path)
        _maybe_upload_after_mutation(full_path, oss_key)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise


@mcp.tool()
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        from excel_mcp.workbook import create_sheet as create_worksheet_impl
        result = create_worksheet_impl(full_path, sheet_name)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise


@mcp.tool()
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = ""
) -> str:
    """Create chart in worksheet."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise


@mcp.tool()
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean"
) -> str:
    """Create pivot table in worksheet."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise


@mcp.tool()
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9"
) -> str:
    """Creates a native Excel table from a specified range of data."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise


@mcp.tool()
def copy_worksheet(
    filepath: str,
    source_sheet: str,
    target_sheet: str
) -> str:
    """Copy worksheet within workbook."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise


@mcp.tool()
def delete_worksheet(
    filepath: str,
    sheet_name: str
) -> str:
    """Delete worksheet from workbook."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = delete_sheet(full_path, sheet_name)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise


@mcp.tool()
def rename_worksheet(
    filepath: str,
    old_name: str,
    new_name: str
) -> str:
    """Rename worksheet in workbook."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise


@mcp.tool()
def get_workbook_metadata(
    filepath: str,
    include_ranges: bool = False
) -> str:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise


@mcp.tool()
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise


@mcp.tool()
def unmerge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Unmerge a range of cells."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise


@mcp.tool()
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet."""
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise


@mcp.tool()
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> str:
    """Copy a range of cells to another location."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        from excel_mcp.sheet import copy_range_operation
        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name  # Use source sheet if target_sheet is None
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise


@mcp.tool()
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up"
) -> str:
    """Delete a range of cells and shift remaining cells."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        from excel_mcp.sheet import delete_range_operation
        result = delete_range_operation(
            full_path,
            sheet_name,
            start_cell,
            end_cell,
            shift_direction
        )
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise


@mcp.tool()
def validate_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted."""
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise


@mcp.tool()
def get_data_validation_info(
    filepath: str,
    sheet_name: str
) -> str:
    """
    Get all data validation rules in a worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet

    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path, _ = _resolve_file_for_operation(filepath)
        from openpyxl import load_workbook
        from excel_mcp.cell_validation import get_all_validation_ranges

        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"

        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()

        if not validations:
            return "No data validation rules found in this worksheet"

        import json
        return json.dumps({
            "sheet_name": sheet_name,
            "validation_rules": validations
        }, indent=2, default=str)

    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise


@mcp.tool()
def insert_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """Insert one or more rows starting at the specified row."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = insert_row(full_path, sheet_name, start_row, count)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise


@mcp.tool()
def insert_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """Insert one or more columns starting at the specified column."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = insert_cols(full_path, sheet_name, start_col, count)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise


@mcp.tool()
def delete_sheet_rows(
    filepath: str,
    sheet_name: str,
    start_row: int,
    count: int = 1
) -> str:
    """Delete one or more rows starting at the specified row."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = delete_rows(full_path, sheet_name, start_row, count)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise


@mcp.tool()
def delete_sheet_columns(
    filepath: str,
    sheet_name: str,
    start_col: int,
    count: int = 1
) -> str:
    """Delete one or more columns starting at the specified column."""
    try:
        full_path, oss_key = _resolve_file_for_operation(filepath)
        result = delete_cols(full_path, sheet_name, start_col, count)
        _maybe_upload_after_mutation(full_path, oss_key)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise


def run_sse():
    """Run Excel MCP server in SSE mode."""
    # Assign value to EXCEL_FILES_PATH in SSE mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with SSE transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="sse")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_streamable_http():
    """Run Excel MCP server in streamable HTTP mode."""
    # Assign value to EXCEL_FILES_PATH in streamable HTTP mode
    global EXCEL_FILES_PATH
    EXCEL_FILES_PATH = os.environ.get("EXCEL_FILES_PATH", "./excel_files")
    # Create directory if it doesn't exist
    os.makedirs(EXCEL_FILES_PATH, exist_ok=True)
    
    try:
        logger.info(f"Starting Excel MCP server with streamable HTTP transport (files directory: {EXCEL_FILES_PATH})")
        mcp.run(transport="streamable-http")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")

def run_stdio():
    """Run Excel MCP server in stdio mode."""
    # No need to assign EXCEL_FILES_PATH in stdio mode
    
    try:
        logger.info("Starting Excel MCP server with stdio transport")
        mcp.run(transport="stdio")
    except KeyboardInterrupt:
        logger.info("Server stopped by user")
    except Exception as e:
        logger.error(f"Server failed: {e}")
        raise
    finally:
        logger.info("Server shutdown complete")